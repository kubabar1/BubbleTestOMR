import os
import cv2
import csv
import openpyxl


def show_contours_cvtColor(image, contours):
    """Show contours detected by cv2.cvtColor

    :param image: image where edges have been detected
    :param contours: contours detected by cv2.cvtColor
    """
    tmp_image = image.copy()
    for contour in contours[0][0]:
        cv2.circle(tmp_image, tuple(contour[0]), radius=3, color=(0, 0, 255), thickness=-1)
    cv2.imshow('cvtColor_contours', tmp_image)
    pass


def show_contours_doc_cnt(image, doc_contour_points):
    """Show contours basis on 4 document contour points

    :param image: image where edges have been detected
    :param doc_contour_points: 4 document contour points
    """
    tmp_image = image.copy()
    a, b, c, d = doc_contour_points[0][0], doc_contour_points[1][0], doc_contour_points[2][0], doc_contour_points[3][0]
    cv2.line(tmp_image, tuple(a), tuple(b), (0, 255, 0), thickness=3, lineType=8)
    cv2.line(tmp_image, tuple(b), tuple(c), (0, 255, 0), thickness=3, lineType=8)
    cv2.line(tmp_image, tuple(c), tuple(d), (0, 255, 0), thickness=3, lineType=8)
    cv2.line(tmp_image, tuple(d), tuple(a), (0, 255, 0), thickness=3, lineType=8)
    cv2.imshow('doc_cnt_contours', tmp_image)
    pass


def show_bounding_box(image, x, y, w, h, frame_name):
    """Show bounding box on image

    :param image: image where bounding box was detected
    :param x: x coordinate of bounding box top left point
    :param y: y coordinate of bounding box top left point
    :param w: width of bounding box
    :param h: height of bounding box
    :param frame_name: name of the frame with visible result
    """
    tmp_image = image.copy()
    cv2.line(tmp_image, (x, y), (x + w, y), (0, 255, 0), thickness=3, lineType=8)
    cv2.line(tmp_image, (x + w, y), (x + w, y + h), (0, 255, 0), thickness=3, lineType=8)
    cv2.line(tmp_image, (x + w, y + h), (x, y + h), (0, 255, 0), thickness=3, lineType=8)
    cv2.line(tmp_image, (x, y + h), (x, y), (0, 255, 0), thickness=3, lineType=8)
    cv2.imshow(frame_name, tmp_image)
    pass


def get_answer_key_default(answers_file_path):
    """Get answer keys from file in default format (correct answers for each question in new line)

    :param answers_file_path: path to file with answers
    :return: answers keys array
    """
    answer_key = {}
    with open(answers_file_path, 'r') as f:
        for idx, line in enumerate(f.readlines()):
            answer_key[idx] = ord(line.replace("\n", "").upper()) - 65
    return answer_key


def get_answer_key_excel(answers_excel_file_path):
    """Get answer keys from file in excel xlsx format (correct answers for each question in new row of excel table in
    first column "A" - each row of table is correct answer to question)

    :param answers_excel_file_path: path to excel file with answers
    :return: answers keys array
    """
    answer_key = {}
    wb_obj = openpyxl.load_workbook(answers_excel_file_path)
    sheet_obj = wb_obj.active
    row_count = sheet_obj.max_row
    for idx in range(1, row_count + 1):
        answer_key[idx - 1] = ord(sheet_obj.cell(row=idx, column=1).value.upper()) - 65
    return answer_key


def get_answer_key_csv(answers_csv_file_path):
    """Get answer keys from file in csv format (correct answers for each question are divided by dash, starting from
    answer on first question, eg. B, E, A, D, B)

    :param answers_csv_file_path: path to csv file with answers
    :return: answers keys array
    """
    answer_key = {}
    with open(answers_csv_file_path) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        row = next(csv_reader)
        for idx, val in enumerate(row):
            answer_key[idx] = ord(val.strip().upper()) - 65
    return answer_key


def save_results(input_images_paths, output_report_dir, scores_array, checked_answers_array):
    """Save results to given directory as file results.xlsx

    :param input_images_paths: path to all input images
    :param output_report_dir: directory where output report with results (report.xlsx) will be generated
    :param scores_array: array containing scores received
    :param checked_answers_array: array with answers checked by user - detected by program basis on input image
    """
    book = openpyxl.Workbook()
    sheet = book.active
    for idx, input_image_path in enumerate(input_images_paths):
        input_file_name = os.path.splitext(os.path.basename(input_image_path))[0]
        sheet.append((input_file_name, scores_array[idx], *[chr(ca + 65) for ca in checked_answers_array[idx]]))
        book.save(os.path.join(output_report_dir, "report.xlsx"))
    pass
